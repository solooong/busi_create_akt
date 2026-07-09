# excel_writer.py
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    Сохраняет список участников в Excel-файл (.xlsx).
    """

    # Стили для красивого вывода
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CELL_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        pass

    def _auto_width(self, worksheet):
        """Автоподбор ширины столбцов по содержимому."""
        for col_cells in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    # Учитываем и русские символы (приблизительно)
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            # Немного запаса
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def save(self,
             participants: List[Dict],
             output_path: Optional[str] = None,
             sheet_name: str = 'Участники') -> str:
        """
        Сохраняет данные в Excel.

        Args:
            participants: список словарей с ключами 'ФИО', 'телефон', 'email', 'доп_данные'
            output_path: полный путь для сохранения. Если None, генерируется имя
                         с датой и временем в текущей рабочей папке.
            sheet_name: название листа в книге.

        Returns:
            str: путь к созданному файлу.
        """
        if not participants:
            logger.warning("Передан пустой список участников. Будет создан пустой файл.")

        # Генерируем имя файла, если не указано
        if output_path is None:
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            default_name = f'participants_{timestamp}.xlsx'
            output_path = str(Path.cwd() / default_name)
            logger.info(f"Путь для сохранения не задан. Используется: {output_path}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Определяем заголовки: основные + дополнительные поля (если есть)
        main_headers = ['ФИО', 'телефон', 'email']
        # Собираем все уникальные ключи из доп_данных (если они хранятся как словари)
        extra_keys = set()
        for p in participants:
            extra = p.get('доп_данные')
            if isinstance(extra, dict):
                extra_keys.update(extra.keys())

        # Сортировка для предсказуемости
        extra_headers = sorted(list(extra_keys))
        all_headers = main_headers + extra_headers

        # Если доп_данные — строка, добавляем колонку "Доп. данные"
        if any(isinstance(p.get('доп_данные'), str) and p.get('доп_данные') for p in participants):
            all_headers = main_headers + ['Доп. данные']
            extra_headers = []  # не будем смешивать

        # Записываем заголовки
        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Записываем данные
        for row_idx, participant in enumerate(participants, 2):
            # Основные поля
            ws.cell(row=row_idx, column=1, value=participant.get('ФИО', '')).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=2, value=participant.get('телефон', '')).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=3, value=participant.get('email', '')).border = self.THIN_BORDER

            extra_data = participant.get('доп_данные', '')
            if extra_headers:
                # Если доп_данные - словарь, раскладываем по колонкам
                if isinstance(extra_data, dict):
                    for col_offset, key in enumerate(extra_headers, 4):
                        value = extra_data.get(key, '')
                        ws.cell(row=row_idx, column=col_offset, value=value).border = self.THIN_BORDER
                # Если строка, но есть колонки extra_headers - выведем строку в первую дополнительную, остальные пустые
                elif isinstance(extra_data, str):
                    ws.cell(row=row_idx, column=4, value=extra_data).border = self.THIN_BORDER
                    for col_offset in range(5, 5 + len(extra_headers) - 1):
                        ws.cell(row=row_idx, column=col_offset, value='').border = self.THIN_BORDER
            elif isinstance(extra_data, str) and extra_data:
                # колонка "Доп. данные"
                ws.cell(row=row_idx, column=4, value=extra_data).border = self.THIN_BORDER

            # Выравнивание
            for col_idx in range(1, len(all_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = self.CELL_ALIGNMENT

        # Автоподбор ширины
        self._auto_width(ws)

        # Сохраняем
        try:
            wb.save(output_path)
            logger.info(f"Данные сохранены в {output_path}")
        except Exception as e:
            logger.error(f"Ошибка сохранения Excel: {e}")
            raise

        return output_path


if __name__ == "__main__":
    # Тестовый запуск
    logging.basicConfig(level=logging.DEBUG)
    test_data = [
        {
            'ФИО': 'Иванов Иван Иванович',
            'телефон': '+7(900)123-45-67',
            'email': 'ivanov@example.com',
            'доп_данные': {'паспорт': '12 34 567890', 'адрес': 'г. Москва, ул. Примерная, д.1'}
        },
        {
            'ФИО': 'Петрова Анна Сергеевна',
            'телефон': '8-912-345-67-89',
            'email': 'petrova@test.ru',
            'доп_данные': 'Семейное положение: замужем'
        }
    ]
    writer = ExcelWriter()
    path = writer.save(test_data)
    print(f"Создан файл: {path}")