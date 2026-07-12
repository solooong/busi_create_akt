# gui_app.py
import os
import sys
import threading
import logging
from tkinter import Checkbutton, Tk, Frame, Label, Entry, Button, Text, Scrollbar, StringVar, BooleanVar
from tkinter import filedialog, ttk, messagebox
from pathlib import Path
from ai_checker import AIChecker
from datetime import datetime
import os

# Получаем текущую дату в формате ГГГГ_ММ_ДД

# Импорты наших модулей
from file_manager import process_folder
from pdf_parser import DDUParser
from excel_writer import ExcelWriter
from docx_filler import DocxFiller

current_date = datetime.now().strftime("%Y_%m_%d")
class LogHandler(logging.Handler):
    """Обработчик логов для вывода в GUI."""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.insert('end', msg + '\n')
        self.text_widget.see('end')


class Application(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Парсер ДДУ и генератор актов")
        self.master.geometry("800x800")
        self.master.resizable(True, True)
        
        # Переменные
        self.folder_path = StringVar(value=os.getcwd())
        self.triggers = StringVar(value="ДДУ, Договор долевого участия,Договор участия в долевом строительстве")
        self.pdf_trigger = StringVar(value="участники долевого строительства")
        self.template_path = StringVar(value="")
        self.running = BooleanVar(value=False)
        self.ai_check = BooleanVar(value=False)
        self.ai_api_url = StringVar(value="http://87.103.253.223:1234/v1/chat/completions")
        self.ai_api_token = StringVar(value="sk-lm-oPeN5xTy:JWPSzgAj1W44Rouiv3qN")
        self.ai_model = StringVar(value="local-model")
        self.create_widgets()
        self.setup_logging()
    
    def create_widgets(self):
        # --- Верхняя панель: настройки ---
        settings_frame = Frame(self.master, padx=10, pady=10)
        settings_frame.pack(fill='x')
        
        # Папка
        Label(settings_frame, text="Корневая папка:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        Entry(settings_frame, textvariable=self.folder_path, width=60).grid(row=0, column=1, padx=5, pady=5)
        Button(settings_frame, text="Обзор...", command=self.select_folder).grid(row=0, column=2, padx=5, pady=5)
        
        # Триггеры файлов
        Label(settings_frame, text="Триггеры файлов (через запятую):").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        Entry(settings_frame, textvariable=self.triggers, width=60).grid(row=1, column=1, padx=5, pady=5)
        
        # Триггер PDF
        Label(settings_frame, text="Триггер в PDF:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        Entry(settings_frame, textvariable=self.pdf_trigger, width=60).grid(row=2, column=1, padx=5, pady=5)
        
        # Шаблон акта
        Label(settings_frame, text="Шаблон акта:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        Entry(settings_frame, textvariable=self.template_path, width=60).grid(row=3, column=1, padx=5, pady=5)
        Button(settings_frame, text="Обзор...", command=self.select_template).grid(row=3, column=2, padx=5, pady=5)

        Checkbutton(settings_frame, text="Проверка через AI (LM Studio)", variable=self.ai_check).grid(
    row=4, column=0, columnspan=3, sticky='w', padx=5, pady=5
)       
        ai_frame = Frame(settings_frame, padx=5, pady=5)
        ai_frame.grid(row=5, column=0, columnspan=3, sticky='ew', pady=5)

        Label(ai_frame, text="API URL:").grid(row=0, column=0, sticky='w', padx=5)
        Entry(ai_frame, textvariable=self.ai_api_url, width=50).grid(row=0, column=1, padx=5)

        Label(ai_frame, text="Токен:").grid(row=1, column=0, sticky='w', padx=5)
        Entry(ai_frame, textvariable=self.ai_api_token, width=50, show="*").grid(row=1, column=1, padx=5)

        Label(ai_frame, text="Модель:").grid(row=2, column=0, sticky='w', padx=5)
        Entry(ai_frame, textvariable=self.ai_model, width=50).grid(row=2, column=1, padx=5)
        # --- Центральная панель: лог ---
        log_frame = Frame(self.master, padx=10, pady=5)
        log_frame.pack(fill='both', expand=True)
        
        Label(log_frame, text="Лог выполнения:").pack(anchor='w')
        
        self.log_text = Text(log_frame, height=20, wrap='word')
        self.log_text.pack(side='left', fill='both', expand=True)
        
        scrollbar = Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.config(yscrollcommand=scrollbar.set)

       
        # self.acts_button.pack(pady=10)
        # --- Нижняя панель: прогресс и кнопка запуска ---
        bottom_frame = Frame(self.master, padx=10, pady=10)
        bottom_frame.pack(fill='x')

        self.progress = ttk.Progressbar(bottom_frame, mode='indeterminate')
        self.progress.pack(fill='x', padx=5, pady=5)

        # Кнопка 1: Парсинг PDF
        self.run_button = Button(
            bottom_frame, 
            text="Запустить обработку PDF", 
            command=self.run_processing,
            bg='#4CAF50',
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=5
        )
        self.run_button.pack(pady=5)

        # Кнопка 2: Только акты из Excel
        self.acts_button = Button(
            bottom_frame,
            text="Заполнить акты из Excel",
            command=self.run_acts_processing,
            bg='#2196F3',
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=5
        )
        self.acts_button.pack(pady=10)
    # Добавьте методы:
    def run_acts_processing(self):
        """Запускает заполнение актов из Excel в отдельном потоке."""
        if self.running.get():
            messagebox.showwarning("Предупреждение", "Обработка уже выполняется")
            return
        
        # Выбор Excel-файла
        excel_path = filedialog.askopenfilename(
            title="Выберите Excel-файл с данными участников",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=self.folder_path.get()
        )
        if not excel_path:
            return
        
        # Выбор шаблона
        template = self.template_path.get()
        if not template or not os.path.exists(template):
            template = filedialog.askopenfilename(
                title="Выберите шаблон акта",
                filetypes=[("Word documents", "*.docx"), ("All files", "*.*")],
                initialdir=self.folder_path.get()
            )
            if not template:
                return
            self.template_path.set(template)
        
        # Выбор папки для сохранения
        output_dir = filedialog.askdirectory(
            title="Выберите папку для сохранения актов",
            initialdir=os.path.join(self.folder_path.get(), "Акты")
        )
        if not output_dir:
            return
        
        self.running.set(True)
        self.run_button.config(state='disabled')
        self.acts_button.config(state='disabled', text="Заполнение...")
        self.progress.start()
        
        thread = threading.Thread(
            target=self._process_acts,
            args=(excel_path, template, output_dir),
            daemon=True
        )
        thread.start()

    def _process_acts(self, excel_path, template_path, output_dir):
        """Обработка заполнения актов в фоновом потоке."""
        logger = logging.getLogger()
        try:
            logger.info("="*60)
            logger.info("ЗАПОЛНЕНИЕ АКТОВ ИЗ EXCEL")
            logger.info("="*60)
            
            from docx_filler import DocxFiller
            import openpyxl
            
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            participants = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                participant = {
                    'ФИО': str(row[0]) if row[0] else '-',
                    'телефон': str(row[1]) if len(row) > 1 and row[1] else '-',
                    'email': str(row[2]) if len(row) > 2 and row[2] else '-',
                    'доп_данные': {}
                }
                
                for i in range(3, len(headers)):
                    if i < len(row) and row[i]:
                        participant['доп_данные'][headers[i]] = str(row[i])
                
                participants.append(participant)
            
            if not participants:
                logger.warning("Нет данных в Excel-файле")
                return
            
            filler = DocxFiller(template_path)
            created = filler.fill_multiple_acts(participants, output_dir)
            
            logger.info(f"✓ Создано {len(created)} актов")
            logger.info("="*60)
            logger.info("ГОТОВО")
            
        except Exception as e:
            logger.error(f"Ошибка: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self.running.set(False)
            self.progress.stop()
            self.run_button.config(state='normal')
            self.acts_button.config(state='normal', text="Заполнить акты из Excel")
    def setup_logging(self):
        """Настраивает логирование в Text виджет."""
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        
        # Удаляем старые хендлеры
        for handler in logger.handlers[:]:
            logger.removeHandler(handler)
        
        # Хендлер для GUI
        gui_handler = LogHandler(self.log_text)
        gui_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(gui_handler)
        
        # Хендлер для файла
        file_handler = logging.FileHandler('processing.log', encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
    
    def select_folder(self):
        """Диалог выбора папки."""
        folder = filedialog.askdirectory(title="Выберите корневую папку", initialdir=self.folder_path.get())
        if folder:
            self.folder_path.set(folder)
    
    def select_template(self):
        """Диалог выбора шаблона акта."""
        file_path = filedialog.askopenfilename(
            title="Выберите шаблон акта",
            filetypes=[("Word документы", "*.docx"), ("Все файлы", "*.*")],
            initialdir=self.folder_path.get()
        )
        if file_path:
            self.template_path.set(file_path)
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке."""
        if self.running.get():
            messagebox.showwarning("Предупреждение", "Обработка уже выполняется")
            return
        
        # Валидация
        folder = self.folder_path.get()
        if not os.path.isdir(folder):
            messagebox.showerror("Ошибка", "Выбранная папка не существует")
            return
        
        self.running.set(True)
        self.run_button.config(state='disabled', text="Обработка...")
        self.progress.start()
        
        # Запуск в отдельном потоке
        thread = threading.Thread(target=self.process_data, daemon=True)
        thread.start()
    
    def process_data(self):
        """Основной процесс обработки данных."""
        logger = logging.getLogger()
        
        try:
            # Шаг 1: поиск файлов
            logger.info("="*60)
            logger.info("НАЧАЛО ОБРАБОТКИ")
            logger.info("="*60)
            
            triggers_set = {t.strip() for t in self.triggers.get().split(',') if t.strip()}
            logger.info(f"Триггеры файлов: {triggers_set}")
            
            pdf_files = process_folder(self.folder_path.get(), triggers=triggers_set)
            pdf_files = [f for f in pdf_files if f.lower().endswith('.pdf')]
            
            if not pdf_files:
                logger.warning("Не найдено PDF-файлов для обработки")
                self.finish_processing()
                return
            
            logger.info(f"Найдено PDF-файлов: {len(pdf_files)}")
            
            # Шаг 2: парсинг PDF
            parser = DDUParser()
            pdf_trigger = self.pdf_trigger.get()
            
            all_participants = []
            seen = set()
            
            participants_with_pdf = []  # список кортежей (участник, путь к pdf)
            for i, pdf_path in enumerate(pdf_files, 1):
                logger.info(f"[{i}/{len(pdf_files)}] Обрабатывается: {os.path.basename(pdf_path)}")
                participants = parser.parse(pdf_path, {'trigger_section': pdf_trigger})
                for p in participants:
                    dop = p.get('доп_данные', {})
                    key = (p.get('ФИО', ''),dop.get('дата_рождения', ''),
                        dop.get('договор_№', ''),dop.get('кадастровый_номер', ''))
                    if key not in seen:
                        seen.add(key)
                        all_participants.append(p)
                        participants_with_pdf.append((p, pdf_path))
                logger.info(f"  Извлечено участников: {len(participants)}")
            
            if not all_participants:
                logger.warning("Ни одного участника не извлечено")
                self.finish_processing()
                return
            
            logger.info(f"Всего уникальных участников: {len(all_participants)}")
            if self.ai_check.get() and all_participants:
                logger.info("Запущена проверка через AI...")
                checker = AIChecker(
                    api_url=self.ai_api_url.get(),
                    api_token=self.ai_api_token.get() or None,
                    model=self.ai_model.get()
                )
                for idx, (p, pdf_path) in enumerate(participants_with_pdf):
                    try:
                        full_text = DDUParser.get_full_text(pdf_path)
                        if full_text:
                            fixed = checker.check_participant(full_text, p)
                            all_participants[idx] = fixed
                            logger.info(f"AI проверил участника: {fixed.get('ФИО', p.get('ФИО', ''))}")
                        else:
                            logger.warning(f"Нет текста для проверки {p.get('ФИО', '')}")
                    except Exception as e:
                        logger.error(f"Ошибка AI для {p.get('ФИО', '')}: {e}")
                logger.info("Проверка AI завершена")
            # Шаг 3: сохранение в Excel
            writer = ExcelWriter()
            excel_path = writer.save(all_participants, os.path.join(self.folder_path.get(), f"participants_{current_date}.xlsx"))
            logger.info(f"Excel сохранён: {excel_path}")
            
            # Шаг 4: заполнение актов
            template = self.template_path.get()
            if template and os.path.exists(template):
                akt_dir = os.path.join(self.folder_path.get(), "Акты")
                os.makedirs(akt_dir, exist_ok=True)
                
                filler = DocxFiller(template)
                created = filler.fill_multiple_acts(all_participants, akt_dir)
                logger.info(f"Создано актов: {len(created)}")
            else:
                logger.info("Шаблон акта не указан или не найден — акты не созданы")
            
            logger.info("="*60)
            logger.info("ОБРАБОТКА УСПЕШНО ЗАВЕРШЕНА")
            logger.info("="*60)
            
        except Exception as e:
            logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            self.finish_processing()
    
    def finish_processing(self):
        """Сбрасывает состояние UI после завершения обработки."""
        self.running.set(False)
        self.progress.stop()
        self.run_button.config(state='normal', text="Запустить обработку")


if __name__ == "__main__":
    root = Tk()
    app = Application(master=root)
    app.mainloop()