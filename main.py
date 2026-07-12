# main.py
import os
import sys
import logging
from pathlib import Path
from datetime import datetime

# Импорты наших модулей
from file_manager import process_folder
from pdf_parser import DDUParser
from excel_writer import ExcelWriter
from docx_filler import DocxFiller

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s',
    handlers=[
        logging.FileHandler("processing.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("main")


def parse_pdfs(root_dir: str, triggers: set = None, pdf_trigger: str = None):
    """
    Режим 1: Парсинг PDF и сохранение в Excel + создание актов.
    
    Args:
        root_dir: корневая папка для обхода
        triggers: множество триггеров для имён файлов
        pdf_trigger: триггер для поиска в тексте PDF
    
    Returns:
        tuple: (путь к Excel, список участников) или (None, [])
    """
    logger.info(f"{'='*60}")
    logger.info(f"РЕЖИМ ПАРСИНГА PDF")
    logger.info(f"{'='*60}")
    logger.info(f"Корневая папка: {root_dir}")
    
    if triggers is None:
        triggers = {"ДДУ", "Договор долевого участия", "Договор участия в долевом строительстве"}
    
    # Шаг 1: поиск и распаковка ZIP
    pdf_files = process_folder(root_dir, triggers=triggers)
    pdf_files = [f for f in pdf_files if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        logger.warning("Не найдено PDF-файлов для обработки.")
        return None, []
    
    logger.info(f"Найдено PDF-файлов: {len(pdf_files)}")
    
    # Шаг 2: парсинг PDF
    parser = DDUParser()
    parser_config = {'trigger_section': pdf_trigger} if pdf_trigger else None
    
    all_participants = []
    seen_participants = set()
    
    for i, pdf_path in enumerate(pdf_files, 1):
        logger.info(f"[{i}/{len(pdf_files)}] Обрабатывается: {os.path.basename(pdf_path)}")
        participants = parser.parse(pdf_path, parser_config)
        
        for p in participants:
            # Уникальный ключ: ФИО + дата рождения
            dop = p.get('доп_данные', {})
            key = (
                p.get('ФИО', ''),
                dop.get('дата_рождения', ''),
                dop.get('договор_№', ''),
                dop.get('кадастровый_номер', '')
            )
            if key not in seen_participants:
                seen_participants.add(key)
                all_participants.append(p)
        
        logger.info(f"  Извлечено участников: {len(participants)}")
    
    if not all_participants:
        logger.warning("Ни одного участника не извлечено.")
        return None, []
    
    logger.info(f"Всего уникальных участников: {len(all_participants)}")
    
    # Шаг 3: сохранение в Excel
    current_date = datetime.now().strftime("%Y_%m_%d")
    excel_path = os.path.join(root_dir, f"participants_{current_date}.xlsx")
    
    writer = ExcelWriter()
    excel_path = writer.save(all_participants, excel_path)
    logger.info(f"Excel сохранён: {excel_path}")
    
    # Шаг 4: создание актов (если есть шаблон)
    template_path = os.path.join(root_dir, "template_akt.docx")
    if os.path.exists(template_path):
        akt_dir = os.path.join(root_dir, "Акты")
        os.makedirs(akt_dir, exist_ok=True)
        
        filler = DocxFiller(template_path)
        created = filler.fill_multiple_acts(all_participants, akt_dir)
        logger.info(f"Создано актов: {len(created)}")
    else:
        logger.info(f"Шаблон акта не найден: {template_path}")
        logger.info("Акты не созданы. Поместите template_akt.docx в корневую папку или укажите путь.")
    
    logger.info(f"{'='*60}")
    logger.info("ОБРАБОТКА ЗАВЕРШЕНА")
    logger.info(f"{'='*60}")
    
    return excel_path, all_participants


def fill_acts_from_excel(excel_path: str, template_path: str, output_dir: str):
    """
    Режим 2: Заполнение актов из готового Excel-файла.
    
    Args:
        excel_path: путь к Excel-файлу с данными
        template_path: путь к шаблону .docx
        output_dir: папка для сохранения актов
    
    Returns:
        list: пути к созданным файлам
    """
    logger.info(f"{'='*60}")
    logger.info(f"РЕЖИМ ЗАПОЛНЕНИЯ АКТОВ")
    logger.info(f"{'='*60}")
    logger.info(f"Excel: {excel_path}")
    logger.info(f"Шаблон: {template_path}")
    logger.info(f"Вывод: {output_dir}")
    
    # Проверки
    if not os.path.exists(excel_path):
        logger.error(f"Excel-файл не найден: {excel_path}")
        return []
    
    if not os.path.exists(template_path):
        logger.error(f"Шаблон не найден: {template_path}")
        return []
    
    # Читаем Excel
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    logger.info(f"Заголовки Excel: {headers}")
    
    participants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # пропускаем пустые строки
            continue
        
        participant = {
            'ФИО': str(row[0]) if row[0] else '-',
            'телефон': str(row[1]) if len(row) > 1 and row[1] else '-',
            'email': str(row[2]) if len(row) > 2 and row[2] else '-',
            'доп_данные': {}
        }
        
        # Остальные колонки → доп.данные
        for i in range(3, len(headers)):
            if i < len(row) and row[i]:
                participant['доп_данные'][headers[i]] = str(row[i])
        
        participants.append(participant)
    
    if not participants:
        logger.warning("Нет данных в Excel-файле")
        return []
    
    logger.info(f"Загружено участников: {len(participants)}")
    
    # Создаём папку вывода
    os.makedirs(output_dir, exist_ok=True)
    
    # Заполняем акты
    filler = DocxFiller(template_path)
    created = filler.fill_multiple_acts(participants, output_dir)
    
    logger.info(f"Создано актов: {len(created)}")
    logger.info(f"{'='*60}")
    logger.info("ГОТОВО")
    logger.info(f"{'='*60}")
    
    return created


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование:")
        print("  python main.py parse [корневая_папка]")
        print("  python main.py acts [excel_файл] [шаблон.docx] [папка_вывода]")
        print()
        print("Примеры:")
        print('  python main.py parse D:\\договоры')
        print('  python main.py acts participants_2026_07_11.xlsx template_akt.docx D:\\акты')
        sys.exit(1)
    
    mode = sys.argv[1].lower()
    
    if mode == 'parse':
        root_folder = sys.argv[2] if len(sys.argv) > 2 else os.getcwd()
        parse_pdfs(root_folder)
    
    elif mode == 'acts':
        if len(sys.argv) < 5:
            print("Укажите: python main.py acts <excel_файл> <шаблон.docx> <папка_вывода>")
            sys.exit(1)
        excel_file = sys.argv[2]
        template_file = sys.argv[3]
        output_folder = sys.argv[4]
        fill_acts_from_excel(excel_file, template_file, output_folder)
    
    else:
        print(f"Неизвестный режим: {mode}")
        print("Доступные режимы: parse, acts")